from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any
from warnings import catch_warnings, simplefilter

from openpyxl import load_workbook

from docifer_backend.config.paths import resolve_project_path


@dataclass(frozen=True)
class GoldenQuestion:
    qa_id: str
    doc_id: str
    category: str
    question: str
    expected_answer: str
    expected_evidence_type: str | None
    expected_pages: str | None
    expected_source_notes: str | None
    should_abstain: bool
    difficulty: str | None
    status: str | None
    reviewer_notes: str | None


@dataclass(frozen=True)
class CorpusDocument:
    doc_id: str
    category: str
    title: str
    source_organization: str
    year: int | None


def load_golden_questions(path: str | Path) -> list[GoldenQuestion]:
    workbook_path = resolve_project_path(path)
    worksheet = _load_sheet(workbook_path, "QA Evaluation Template")
    headers = [cell.value for cell in worksheet[1]]
    questions: list[GoldenQuestion] = []

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        if not record.get("QA ID") or not record.get("Question"):
            continue
        questions.append(
            GoldenQuestion(
                qa_id=str(record["QA ID"]),
                doc_id=str(record["Doc ID"]),
                category=str(record["Question Category"]),
                question=str(record["Question"]),
                expected_answer=str(record.get("Expected Answer") or ""),
                expected_evidence_type=_optional_str(record.get("Expected Evidence Type")),
                expected_pages=_optional_str(record.get("Expected Page(s)")),
                expected_source_notes=_optional_str(record.get("Expected Source Notes")),
                should_abstain=_parse_bool(record.get("Should Abstain?")),
                difficulty=_optional_str(record.get("Difficulty")),
                status=_optional_str(record.get("Status")),
                reviewer_notes=_optional_str(record.get("Reviewer Notes")),
            )
        )
    return questions


def load_corpus_documents(path: str | Path) -> dict[str, CorpusDocument]:
    workbook_path = resolve_project_path(path)
    worksheet = _load_sheet(workbook_path, "Starter Corpus")
    headers = [cell.value for cell in worksheet[1]]
    documents: dict[str, CorpusDocument] = {}

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        if not record.get("Doc ID"):
            continue
        documents[str(record["Doc ID"])] = CorpusDocument(
            doc_id=str(record["Doc ID"]),
            category=str(record.get("Category") or ""),
            title=str(record.get("Document Title") or ""),
            source_organization=str(record.get("Source Organization") or ""),
            year=_optional_int(record.get("Year")),
        )
    return documents


def _load_sheet(path: Path, sheet_name: str):
    with catch_warnings():
        simplefilter("ignore")
        workbook = load_workbook(path, data_only=True)
    return workbook[sheet_name]


def _optional_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _optional_int(value: Any) -> int | None:
    if value is None:
        return None
    return int(value)


def _parse_bool(value: Any) -> bool:
    return str(value or "").strip().lower() in {"yes", "true", "1"}
